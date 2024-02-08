from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
import nltk
nltk.download(['punkt'])


from datetime import datetime

path = r"C:\Users\PC\PycharmProjects\pythonProject1\Sumit-Project\chromedriver.exe"
driver = webdriver.Chrome(path)
driver.get("https://google.com")

today = datetime.today().strftime('%A').lower()
print(today)

book = load_workbook("Worksheet.xlsx")

sheet = book[today]

value_range = sheet['D5':'D14']

for row_index, row in enumerate(value_range, start=5):
    for cell_index, cell in enumerate(row, start=6):
        key = cell.value
        search = driver.find_element("name", "q")
        search.send_keys(key)
        time.sleep(5)
        try:
            main_s = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "G43f7e"))
            )
            search_result = main_s.text
            sentences = nltk.sent_tokenize(search_result)
            print("Total searches=", sentences)

            sentences = [sentence.strip() for sentence in sentences[0].split('\n')]
            longest_sentence = max(sentences, key=len)
            shortest_sentence = min(sentences, key=len)

            print("Longest sentence:", longest_sentence)
            print("Shortest sentence:", shortest_sentence)


            sheet.cell(row=row_index, column=5).value = longest_sentence

            sheet.cell(row=row_index, column=6).value = shortest_sentence

            print("Updated value in Excel - Longest sentence:", sheet.cell(row=row_index, column=6).value)
            print("Updated value in Excel - Shortest sentence:", sheet.cell(row=row_index, column=7).value)

        except Exception as e:
            print("An error occurred:", e)
        driver.find_element("name", "q").clear()
        time.sleep(2)

book.save("Worksheet.xlsx")
driver.quit()
